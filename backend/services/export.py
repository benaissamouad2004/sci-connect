# SERVICE: Génération Excel 4 feuilles — openpyxl
# EDITABLE: modifier les colonnes, couleurs et seuils ci-dessous
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference


# ═══ CONSTANTES COULEURS ═══
# Changer ces valeurs pour modifier toutes les couleurs Excel d'un coup
COLOR_HEADER_BG   = '005F54'   # Teal primaire — en-tête de toutes les feuilles
COLOR_HEADER_TEXT = 'FFFFFF'   # Blanc — texte en-tête
COLOR_ROW_EVEN    = 'F7F6F2'   # Fond lignes paires
COLOR_ROW_ODD     = 'FFFFFF'   # Fond lignes impaires
COLOR_ACCENT      = 'C9A84C'   # Or — mise en évidence
COLOR_DANGER      = 'C0392B'   # Rouge — problèmes qualité
COLOR_BORDER      = 'E2DED6'   # Gris clair — bordures


def _header_style():
    return (
        PatternFill(start_color=COLOR_HEADER_BG, end_color=COLOR_HEADER_BG, fill_type='solid'),
        Font(color=COLOR_HEADER_TEXT, bold=True, name='Calibri', size=11),
        Alignment(horizontal='center', vertical='center', wrap_text=True),
    )

def _apply_header(ws, headers, start_row=1):
    fill, font, align = _header_style()
    ws.append(headers)
    row_num = ws.max_row
    for cell in ws[row_num]:
        cell.fill      = fill
        cell.font      = font
        cell.alignment = align
    ws.row_dimensions[row_num].height = 28
    ws.freeze_panes = f'A{row_num + 1}'

def _alternating_rows(ws, start_row=2):
    fill_e = PatternFill(start_color=COLOR_ROW_EVEN, end_color=COLOR_ROW_EVEN, fill_type='solid')
    fill_o = PatternFill(start_color=COLOR_ROW_ODD,  end_color=COLOR_ROW_ODD,  fill_type='solid')
    data_font = Font(name='Calibri', size=10)
    thin  = Side(style='thin', color=COLOR_BORDER)
    brd   = Border(bottom=thin)
    for i, row in enumerate(ws.iter_rows(min_row=start_row), start=start_row):
        fill = fill_e if i % 2 == 0 else fill_o
        for cell in row:
            cell.fill      = fill
            cell.font      = data_font
            cell.border    = brd
            cell.alignment = Alignment(vertical='center')

def _autofit(ws):
    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=8) + 4
        ws.column_dimensions[col[0].column_letter].width = min(max_len, 50)


def generate_excel(questionnaire, responses, schools_data=None):
    """
    Génère un fichier Excel avec 4 feuilles :
      1. Données brutes
      2. Statistiques auto
      3. Profils démographiques
      4. Fiabilité & Qualité
    EDITABLE: modifier les colonnes dans chaque feuille ci-dessous
    """
    # Index école → nom lisible depuis schools.json
    school_names = {}
    uni_names    = {}
    if schools_data:
        for uni in schools_data.get('universities', []):
            uni_names[uni['id']] = uni['name']
            for s in uni.get('schools', []):
                school_names[s['id']] = s['name']

    # Enrichir les réponses avec les données utilisateur
    enriched = []
    for r in responses:
        user = _get_respondent_user(r)
        enriched.append({
            'response':    r,
            'user':        user,
            'school_name': school_names.get(user.school_id, '—') if user else '—',
            'uni_name':    uni_names.get(user.university_id, '—') if user else '—',
            'domain':      getattr(user, 'domains', None),
            'level':       getattr(user, 'level', '—') if user else '—',
        })

    wb = Workbook()

    # ════ FEUILLE 1 — Données brutes ════
    _build_sheet1(wb, questionnaire, enriched)

    # ════ FEUILLE 2 — Statistiques auto ════
    _build_sheet2(wb, questionnaire, enriched)

    # ════ FEUILLE 3 — Profils démographiques ════
    _build_sheet3(wb, questionnaire, enriched, uni_names)

    # ════ FEUILLE 4 — Fiabilité & Qualité ════
    _build_sheet4(wb, questionnaire, enriched)

    # Supprimer la feuille par défaut vide si elle existe
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _build_sheet1(wb, q, enriched):
    """
    Feuille 1 : Données brutes
    EDITABLE: ajouter/supprimer des colonnes dans headers ci-dessous
    Données confidentielles — usage interne uniquement
    """
    ws = wb.active
    ws.title = 'Données brutes'

    # Ligne de confidentialité en haut
    ws.append(['⚠ Données confidentielles — usage interne uniquement — ne pas diffuser'])
    conf_cell = ws['A1']
    conf_cell.font = Font(bold=True, color='922B21', name='Calibri', size=10)
    conf_cell.fill = PatternFill(start_color='FEF3F2', end_color='FEF3F2', fill_type='solid')
    ws.merge_cells('A1:J1')
    ws.row_dimensions[1].height = 20

    # EDITABLE: colonnes feuille 1
    # Email académique : affiché uniquement pour les étudiants vérifiés (respondent_type == 'verified')
    headers = ['#', 'Email académique', 'Type', 'Université', 'École',
               'Filière', 'Niveau', 'Date', 'Complétion (%)', 'Statut validation']
    _apply_header(ws, headers, start_row=2)

    for i, e in enumerate(enriched, 1):
        r      = e['response']
        domain = (e['domain'] or [None])[0] if isinstance(e['domain'], list) else (e['domain'] or '—')
        # Email uniquement si répondant vérifié — anonyme sinon
        email  = r.respondent_email if r.respondent_type == 'verified' else 'Anonyme'
        ws.append([
            i,
            email,
            _format_type(r.respondent_type),
            e['uni_name'],
            e['school_name'],
            domain or '—',
            e['level'] or '—',
            r.created_at.strftime('%d/%m/%Y %H:%M') if r.created_at else '—',
            f"{r.completion_percentage:.0f}%" if r.completion_percentage is not None else '—',
            'Validé ✓' if r.validated_by_emitter else ('Ignoré' if r.ignored_by_emitter else 'En attente'),
        ])

    _alternating_rows(ws, start_row=3)
    _autofit(ws)


def _build_sheet2(wb, q, enriched):
    """
    Feuille 2 : Statistiques automatiques
    Résumé par type de répondant
    EDITABLE: ajouter des métriques supplémentaires ici
    """
    ws = wb.create_sheet('Statistiques auto')

    total    = len(enriched)
    verified = sum(1 for e in enriched if e['response'].respondent_type == 'verified')
    public_  = sum(1 for e in enriched if e['response'].respondent_type == 'public')
    anon     = total - verified - public_
    validated = sum(1 for e in enriched if e['response'].validated_by_emitter)
    avg_compl = (sum((e['response'].completion_percentage or 0) for e in enriched) / total) if total else 0

    # En-tête de titre
    fill, font, align = _header_style()
    ws['A1'] = f'Statistiques — {q.title}'
    ws['A1'].fill  = fill
    ws['A1'].font  = font
    ws.merge_cells('A1:C1')
    ws.row_dimensions[1].height = 28

    # EDITABLE: métriques résumé feuille 2
    summary = [
        ['Métrique', 'Valeur', 'Note'],
        ['Total réponses',           total,              ''],
        ['Étudiants vérifiés',       verified,           f'{round(verified/total*100, 1) if total else 0}%'],
        ['Public général',           public_,            f'{round(public_/total*100, 1) if total else 0}%'],
        ['Anonymes',                 anon,               f'{round(anon/total*100, 1) if total else 0}%'],
        ['Validées par émetteur',    validated,          f'{round(validated/total*100, 1) if total else 0}%'],
        ['Complétion moyenne',       f'{avg_compl:.1f}%', ''],
        ['Objectif réponses',        q.target_count or 100, ''],
        ['Progression objectif',     f'{round(total/(q.target_count or 100)*100, 1)}%', ''],
    ]

    fill_h, font_h, align_h = _header_style()
    for row_idx, row_data in enumerate(summary, start=2):
        ws.append(row_data)
        if row_idx == 2:  # ligne d'en-tête du tableau
            for cell in ws[row_idx]:
                cell.fill  = fill_h
                cell.font  = font_h
                cell.alignment = align_h

    _alternating_rows(ws, start_row=3)

    # Graphique camembert types de répondants
    if total > 0:
        data_ref = Reference(ws, min_col=2, min_row=3, max_row=5)
        cats_ref = Reference(ws, min_col=1, min_row=3, max_row=5)
        pie = PieChart()
        pie.title  = 'Répartition par type'
        pie.add_data(data_ref)
        pie.set_categories(cats_ref)
        pie.width  = 15
        pie.height = 12
        ws.add_chart(pie, 'E2')

    _autofit(ws)


def _build_sheet3(wb, q, enriched, uni_names):
    """
    Feuille 3 : Profils démographiques
    Répartition par université, niveau et domaine
    EDITABLE: ajouter d'autres axes démographiques ici
    """
    ws = wb.create_sheet('Profils démographiques')

    fill, font, align = _header_style()
    ws['A1'] = 'Profils démographiques'
    ws['A1'].fill  = fill
    ws['A1'].font  = font
    ws.merge_cells('A1:C1')
    ws.row_dimensions[1].height = 28

    # ─ Par université
    uni_counts = {}
    for e in enriched:
        key = e['uni_name'] or '—'
        uni_counts[key] = uni_counts.get(key, 0) + 1

    row = 3
    ws.cell(row, 1, 'Université').fill  = fill
    ws.cell(row, 1, 'Université').font  = font
    ws.cell(row, 2, 'Réponses').fill   = fill
    ws.cell(row, 2, 'Réponses').font   = font
    ws.cell(row, 3, '%').fill          = fill
    ws.cell(row, 3, '%').font          = font
    row += 1
    total = len(enriched) or 1
    for name, cnt in sorted(uni_counts.items(), key=lambda x: -x[1]):
        ws.cell(row, 1, name)
        ws.cell(row, 2, cnt)
        ws.cell(row, 3, f'{cnt/total*100:.1f}%')
        row += 1

    # ─ Par niveau
    row += 1
    ws.cell(row, 1, 'Niveau').fill   = fill
    ws.cell(row, 1, 'Niveau').font   = font
    ws.cell(row, 2, 'Réponses').fill = fill
    ws.cell(row, 2, 'Réponses').font = font
    row += 1
    level_counts = {}
    for e in enriched:
        key = e['level'] or '—'
        level_counts[key] = level_counts.get(key, 0) + 1
    for name, cnt in sorted(level_counts.items(), key=lambda x: -x[1]):
        ws.cell(row, 1, name)
        ws.cell(row, 2, cnt)
        row += 1

    # Graphique barres universités
    if uni_counts:
        start_r = 4
        end_r   = start_r + len(uni_counts) - 1
        data_r  = Reference(ws, min_col=2, min_row=start_r, max_row=end_r)
        cats_r  = Reference(ws, min_col=1, min_row=start_r, max_row=end_r)
        bar = BarChart()
        bar.title  = 'Réponses par université'
        bar.add_data(data_r)
        bar.set_categories(cats_r)
        bar.width  = 18
        bar.height = 12
        ws.add_chart(bar, 'E3')

    _autofit(ws)


def _build_sheet4(wb, q, enriched):
    """
    Feuille 4 : Fiabilité & Qualité
    EDITABLE: le seuil suspect est configurable dans routes/responses.py (MIN_DURATION_SUSPECT)
    """
    ws = wb.create_sheet('Fiabilité & Qualité')

    fill, font, align = _header_style()
    ws['A1'] = 'Fiabilité & Qualité des réponses'
    ws['A1'].fill  = fill
    ws['A1'].font  = font
    ws.merge_cells('A1:D1')
    ws.row_dimensions[1].height = 28

    total     = len(enriched)
    complete  = sum(1 for e in enriched if e['response'].is_complete)
    abandoned = total - complete
    validated = sum(1 for e in enriched if e['response'].validated_by_emitter)
    ignored   = sum(1 for e in enriched if e['response'].ignored_by_emitter)
    suspect_count = sum(1 for e in enriched if e['response'].is_suspect or False)

    summary = [
        ['Indicateur',               'Valeur',  '%',       'Interprétation'],
        ['Total réponses',           total,      '100%',    ''],
        ['Réponses complètes',       complete,   f'{complete/total*100:.1f}%' if total else '—', 'Idéal > 80%'],
        ['Réponses abandonnées',     abandoned,  f'{abandoned/total*100:.1f}%' if total else '—', 'Alerte si > 30%'],
        ['Validées par émetteur',    validated,  f'{validated/total*100:.1f}%' if total else '—', ''],
        ['Ignorées par émetteur',    ignored,    f'{ignored/total*100:.1f}%' if total else '—', ''],
        ['Réponses suspectes',       suspect_count, f'{suspect_count/total*100:.1f}%' if total else '—', 'Durée < 60s — 0 pts attribués'],
    ]

    fill_h, font_h, _ = _header_style()
    for row_idx, row_data in enumerate(summary, start=2):
        ws.append(row_data)
        if row_idx == 2:
            for cell in ws[row_idx]:
                cell.fill = fill_h
                cell.font = font_h

    _alternating_rows(ws, start_row=3)

    # ─ Répondants suspects (durée trop courte, détectés par le backend)
    # EDITABLE: la détection suspect est faite dans routes/responses.py (seuil 60s)
    suspects = [e for e in enriched if (e['response'].is_suspect or False)]

    row = 10
    ws.cell(row, 1, f'Répondants suspects ({len(suspects)} — durée < 60s)')
    ws.cell(row, 1).fill = PatternFill(start_color=COLOR_DANGER, end_color=COLOR_DANGER, fill_type='solid')
    ws.cell(row, 1).font = Font(color='FFFFFF', bold=True, name='Calibri', size=11)
    ws.merge_cells(f'A{row}:D{row}')
    row += 1

    if suspects:
        hdrs = ['Type', 'Durée (sec)', 'Date', 'Note']
        ws.append(hdrs)
        for cell in ws[row]:
            cell.fill = fill_h
            cell.font = font_h
        row += 1
        danger_fill = PatternFill(start_color='FEF3F2', end_color='FEF3F2', fill_type='solid')
        for e in suspects:
            r = e['response']
            ws.append([
                _format_type(r.respondent_type),
                r.duration_seconds or '—',
                r.created_at.strftime('%d/%m/%Y %H:%M') if r.created_at else '—',
                'Durée inférieure au minimum requis (60s)',
            ])
            for cell in ws[ws.max_row]:
                cell.fill = danger_fill
    else:
        ws.cell(row, 1, 'Aucun répondant suspect détecté ✓')
        ws.cell(row, 1).font = Font(color='1A7A5E', bold=True, name='Calibri')

    _autofit(ws)


def _format_type(respondent_type):
    mapping = {
        'verified':  'Étudiant vérifié ✓',
        'public':    'Public général',
        'anonymous': 'Anonyme',
    }
    return mapping.get(respondent_type, respondent_type or '—')


def _get_respondent_user(response):
    try:
        if not response.respondent_id:
            return None
        from backend.models import User
        return User.query.filter_by(id=response.respondent_id).first()
    except Exception:
        return None